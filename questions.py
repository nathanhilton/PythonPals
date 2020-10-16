import random
import Button
from openpyxl import load_workbook

wk = load_workbook(filename='python_questions.xlsx')
ws = wk.active
    
def load_question(category):
    print("Question Categories: ")
    print("1 Syntax\n2 Vocabulary\n3 Logic\n4 Number Conversion\n5 General\n")
    value = 0
    
    if category == '1':
        value = random.randrange(2, 7)
    elif category == '2':
        value = random.randrange(7, 12)
    elif category == '3':
        value = random.randrange(12, 17)
    elif category == '4':
        value = random.randrange(17, 22)
    else:
        value = random.randrange(2, 22)

    return value

def get_question(value):
    question = ws['B' + str(value)].value
    #print(ws[question].value)
    choices = ws['D' + str(value)].value
    #print(ws[choices].value)
    answer = ws['E' + str(value)].value

    return question, choices, answer 

def get_result(guess, answer):
    if guess == answer or guess == answer.capitalize() or guess == answer + '.' or guess == answer.capitalize() + '.':
        #print("Correct!")
        return True
    else:
        #print("Incorrect")
        return False




# Press the green button in the gutter to run the script.
if __name__ == '__main__':
    load_question()
