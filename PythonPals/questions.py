import random

from openpyxl import load_workbook, Workbook


wk = load_workbook(filename="python_questions.xlsx")
ws = wk.active


def load_question(category):
    print("Question Categories: ")
    print("1 Syntax\n2 Vocabulary\n3 Logic\n4 Number Conversion\n5 General\n")
    value = 0

    if category == int(1):
        value = random.randrange(2, 7)
    elif category == int(2):
        value = random.randrange(7, 12)
    elif category == int(3):
        value = random.randrange(12, 17)
    elif category == (4):
        value = random.randrange(17, 22)
    else:
        value = random.randrange(2, 22)

    print(value)
    return value


'''
def get_question(value):
    question = ws['B' + str(value)].value
    #print(ws[question].value)
    choices = ws['D' + str(value)].value
    #print(ws[choices].value)
    answer = ws['E' + str(value)].value

    return question, choices, answer 
'''


def get_result(guess, answer):
    if guess == answer or guess == answer.capitalize() or guess == answer + '.' or guess == answer.capitalize() + '.':
        # print("Correct!")
        return True
    else:
        # print("Incorrect")
        return False


def get_question(value):
    # list returned has the info
    # 0 has the question
    # 1 has answer A
    # 2 has answer B
    # 3 has answer C
    # 4 has answer D
    # 5 has the correct answer (ex "A")
    the_question = ["", "", "", "", "", ""]

    question = 'B' + str(value)
    the_question[0] = ws[question].value

    choiceA = 'F' + str(value)
    the_question[1] = ws[choiceA].value

    choiceB = 'G' + str(value)
    the_question[2] = ws[choiceB].value

    choiceC = 'H' + str(value)
    the_question[3] = ws[choiceC].value

    choiceD = 'I' + str(value)
    the_question[4] = ws[choiceD].value

    the_question[5] = ws['E' + str(value)].value

    return the_question


def changeQuestionDeck(filename):
    wk = load_workbook(filename)
    ws = wk.active
    print("Loaded")







# Press the green button in the gutter to run the script.
#if __name__ == '__main__':
    #load_question()
