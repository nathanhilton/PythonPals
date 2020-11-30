# import Button
# import questions
import pygame
import pygame.gfxdraw
# from player import Player, Coffee, Ruby, Eye
import random
from openpyxl import load_workbook, Workbook
from pygame.locals import *

class Screen():
    def __init__(self, width, height):
        self.width = width
        self.height = height

class damage():
    def __init__(self, userDam, enemyDam):
        self.userDamage = userDam
        self.enemyDamage = enemyDam

    def modify(self, newUserDam, newEnemyDam):
        self.userDamage = newUserDam
        self.enemyDamage = newEnemyDam

class music():
    def __init__(self):
        self.play = True
        self.name = "elf"
    def mute(self):
        self.play = False
    def unmute(self):
        self.play = True
    def change(self, new_name):
        self.name = new_name

#-----from player.py-----#
class Player(pygame.sprite.Sprite):
    def __init__(self):
        pygame.sprite.Sprite.__init__(self)
        self.images = []

        self.images.append(pygame.image.load('snake1.png'))
        self.images.append(pygame.image.load('Snake_start.png'))
        self.images.append(pygame.image.load('Snake_Firemid.png'))
        self.images.append(pygame.image.load('snake2.png'))
        self.images.append(pygame.image.load('Snake_fire_end.png'))
        self.images.append(pygame.image.load('snake4.png'))
        self.images.append(pygame.image.load('snake6.png'))

        self.index = 0
        self.image = self.images[self.index]
        # self.rect = pygame.Rect(250, 575, 600, 600)
        self.rect = pygame.Rect(theScreen.width / 5, theScreen.height / 1.6, 600, 600)

    # def resize(self, scale):
    #     self.height = self.height * scale
    #     self.width = self.width * scale
    #     self.image = pygame.transform.rotozoom(self.images[self.index], 0, scale)

    def resize(self):
        self.rect = pygame.Rect(theScreen.width/5, theScreen.height/1.6, 600, 600)
        # self.image = pygame.transform.rotozoom(self.images[self.index], 0, theScreen.width/200)

    def update(self):
        self.index += 1
        if self.index >= len(self.images):
            self.index = 0
        self.image = self.images[self.index]

    # def changeLocation(self, x, y):
    #     self.x = x
    #     self.y = y
    #     self.rect = pygame.Rect(x, y, self.width, self.height)

class Coffee(pygame.sprite.Sprite):
    def __init__(self):
        pygame.sprite.Sprite.__init__(self)
        self.images = []
        self.images.append(pygame.image.load('coffee1.png'))
        self.images.append(pygame.image.load('coffee2.png'))
        self.images.append(pygame.image.load('coffee3.png'))
        self.images.append(pygame.image.load('coffee4.png'))
        self.images.append(pygame.image.load('coffee5.png'))
        self.images.append(pygame.image.load('coffee6.png'))

        self.index = 0
        self.image = self.images[self.index]
        # self.rect = pygame.Rect(1000, 450, 500, 500)
        self.rect = pygame.Rect(theScreen.width/1.5, theScreen.height/2, 600, 600)

    def update(self):
        self.index += 1
        if self.index >= len(self.images):
            self.index = 0
        self.image = self.images[self.index]

    def resize(self):
        self.rect = pygame.Rect(theScreen.width/1.5, theScreen.height/2, 600, 600)

    # def resize(self, scale):
    #     self.height = self.height * scale
    #     self.width = self.width * scale
    #     self.image = pygame.transform.rotozoom(self.images[self.index], 0, scale)
    #
    # def changeLocation(self, x, y):
    #     self.x = x
    #     self.y = y
    #     self.rect = pygame.Rect(x, y, self.width, self.height)

class Ruby(pygame.sprite.Sprite):
    def __init__(self):
        pygame.sprite.Sprite.__init__(self)
        self.images = []
        self.images.append(pygame.image.load('Ruby_idle.png'))
        self.images.append(pygame.image.load('Ruby_left.png'))
        self.images.append(pygame.image.load('Ruby_prep.png'))
        self.images.append(pygame.image.load('Ruby_start.png'))
        self.images.append(pygame.image.load('Ruby_fire.png'))
        self.images.append(pygame.image.load('Ruby_start.png'))
        self.images.append(pygame.image.load('Ruby_prep.png'))
        self.images.append(pygame.image.load('Ruby_left.png'))
        self.images.append(pygame.image.load('Ruby_hurt.png'))

        self.index = 0
        self.image = self.images[self.index]
        # self.rect = pygame.Rect(800, 350, 500, 500)
        self.rect = pygame.Rect(theScreen.width / 2, theScreen.height / 2.5, 600, 600)

    def update(self):
        self.index += 1
        if self.index >= len(self.images):
            self.index = 0
        self.image = self.images[self.index]

    def resize(self):
        self.rect = pygame.Rect(theScreen.width/2, theScreen.height/2.5, 600, 600)

    # def resize(self, scale):
    #     self.height = self.height * scale
    #     self.width = self.width * scale
    #     self.image = pygame.transform.rotozoom(self.images[self.index], 0, scale)

    # def changeLocation(self, x, y):
    #     self.x = x
    #     self.y = y
    #     self.rect = pygame.Rect(x, y, self.width, self.height)

class Eye(pygame.sprite.Sprite):
    def __init__(self):
        pygame.sprite.Sprite.__init__(self)
        self.images = []
        self.images.append(pygame.image.load('C1.png'))
        self.images.append(pygame.image.load('C2.png'))
        self.images.append(pygame.image.load('C3.png'))
        self.images.append(pygame.image.load('C4.png'))


        self.index = 0
        self.image = self.images[self.index]
        # self.rect = pygame.Rect(850, 400, 500, 500)
        self.rect = pygame.Rect(theScreen.width/2, theScreen.height/2.5, 600, 600)

    def update(self):
        self.index += 1
        if self.index >= len(self.images):
            self.index = 0
        self.image = self.images[self.index]

    def resize(self):
        self.rect = pygame.Rect(theScreen.width/2, theScreen.height/2.5, 600, 600)

    # def resize(self, scale):
    #     self.height = self.height * scale
    #     self.width = self.width * scale
    #     self.image = pygame.transform.rotozoom(self.images[self.index], 0, scale)
    #
    # def changeLocation(self, x, y):
    #     self.x = x
    #     self.y = y
    #     self.rect = pygame.Rect(x, y, self.width, self.height)

#-----from questions.py-----#
def initialize(filename):
    wk = load_workbook(filename=filename)

    ws = wk.active
    return ws

def load_question(category):
    print("Question Categories: ")
    print("1 Syntax\n2 Vocabulary\n3 Logic\n4 Number Conversion\n5 General\n")
    value = 0

    if category == int(1):
        value = random.randrange(2, 7)
    elif category == int(2):
        value = random.randrange(7, 12)
    elif category == int(3):
        value = random.randrange(12, 17)
    elif category == int(4):
        value = random.randrange(17, 22)
    else:
        value = random.randrange(2, 22)

    print(value)
    return value

def get_result(guess, answer):
    if guess == answer or guess == answer.capitalize() or guess == answer + '.' or guess == answer.capitalize() + '.':
        # print("Correct!")
        return True
    else:
        # print("Incorrect")
        return False

def get_question(value,ws):
    # list returned has the info
    # 0 has the question
    # 1 has answer A
    # 2 has answer B
    # 3 has answer C
    # 4 has answer D
    # 5 has the correct answer (ex "A")
    the_question = ["", "", "", "", "", ""]

    question = 'B' + str(value)
    the_question[0] = ws[question].value

    choiceA = 'F' + str(value)
    the_question[1] = ws[choiceA].value

    choiceB = 'G' + str(value)
    the_question[2] = ws[choiceB].value

    choiceC = 'H' + str(value)
    the_question[3] = ws[choiceC].value

    choiceD = 'I' + str(value)
    the_question[4] = ws[choiceD].value

    the_question[5] = ws['E' + str(value)].value

    return the_question

def changeQuestionDeck(filename, ws):
    wk = load_workbook(filename=filename)
    ws = wk.active
    print("Loaded")
    return ws

#-----from Button.py-----#
class button():
    def __init__(self, color, x, y, width, height, text=''):
        self.color = color
        self.x = x
        self.y = y
        self.width = width
        self.height = height
        self.text = text

    def modify(self, x, y, width, height):
        self.x = x
        self.y = y
        self.width = width
        self.height = height

    def draw(self, screen, fontSize, text_color, center=False, outline=None):
        # Call this method to draw the button on the screen
        if outline:
            pygame.draw.rect(screen, outline, (round(self.x - 2, self.y - 2), round(self.width + 4, self.height + 4)),
                             0)

        pygame.draw.rect(screen, self.color, (round(self.x), round(self.y), round(self.width), round(self.height)), 0)

        if self.text != '':
            font = pygame.font.Font('JandaManateeSolid.ttf', fontSize)
            text = font.render(self.text, 1, text_color)
            if center:
                screen.blit(text, (round(self.x + (self.width / 2 - text.get_width() / 2)),
                                   round(self.y + (self.height / 2 - text.get_height() / 2))))
            else:
                screen.blit(text, (round(self.x + self.width), round(self.y)))

        pygame.display.update()

    def isOver(self, pos):
        if pos[0] > self.x and pos[0] < self.x + self.width:
            if pos[1] > self.y and pos[1] < self.y + self.height:
                return True

        return False

class text():
    def __init__(self, color, x, y, width, height, textSize, text):
        self.color = color
        self.x = x
        self.y = y
        self.width = width
        self.height = height
        self.text = text
        self.textSize = textSize

    def modify(self, x, y, width, height):
        self.x = x
        self.y = y
        self.width = width
        self.height = height

    def draw(self, screen, fontSize, center):
        #pygame.draw.rect(screen, self.color, (round(self.x), round(self.y), round(self.width), round(self.height)), 0)

        if len(self.text) != 0:
            size = len(self.text)
            font = pygame.font.Font("JandaManateeSolid.ttf", fontSize)
            i = 0
            if center:
                for part in self.text:
                    text = font.render(part, True, self.color)
                    screen.blit(text, (round(self.x + (self.width / 2 - text.get_width() / 2)),
                                   round(self.y + (self.height / 2 - text.get_height() / 2) + i)))
                    i += 37
            else:
                for part in self.text:
                    text = font.render(part, True, self.color)
                    screen.blit(text, (round(self.x + self.width), round(self.y + i)))
                    i += 37

class healthBar():
    def __init__(self, x, y, width, height, health, orientation):
        self.x = x
        self.y = y
        self.width = width
        self.height = height
        self.health = health
        self.orientation = orientation

    def modify(self, x, y, width, height):
        self.x = x
        self.y = y
        self.width = width
        self.height = height

    def set_health(self, newHealth):
        self.health = newHealth

    def draw(self, screen):
        pygame.draw.rect(screen, (255, 255, 255), (self.x - 2, self.y - 2, self.width + 4, self.height + 4))
        pygame.draw.rect(screen, (0, 0, 0), (self.x, self.y, self.width, self.height))
        if self.health != 0:
            if self.health >= 55:
                pygame.draw.rect(screen, (0, 128, 0),
                                 (self.x, self.y, self.width - (self.width * ((100 - self.health) / 100)), self.height))
            elif self.health < 25:
                pygame.draw.rect(screen, (255, 0, 0),
                                 (self.x, self.y, self.width - (self.width * ((100 - self.health) / 100)), self.height))
            else:
                pygame.draw.rect(screen, (255, 255, 0),
                                 (self.x, self.y, self.width - (self.width * ((100 - self.health) / 100)), self.height))
            # print(self.health)
            # font = pygame.font.SysFont('comicsans', 40)
            # text = font.render("HP:", 1, (0, 0, 0), (204,204,0))
            # text_rect = text.get_rect()
            # text_rect.center = (self.x + (self.width * 0.15), self.y + (self.height * 0.5))
            # screen.blit(text, text_rect)
        pygame.display.update()

#-----initialize objects etc-----#
ws = initialize("python_questions.xlsx")
screen = pygame.display.set_mode((0,0), pygame.RESIZABLE)
white = (255, 255, 255)
purple = (110, 113, 198)
blue = (64, 127, 194)
indigo = (51, 57, 255)
color_light = (170, 170, 170)
color_dark = (100, 100, 100)
fuschia = (255, 0, 255)
black = (0, 0, 0)
gold = (197, 179, 88)
lime = (153, 255, 51)
width = screen.get_width()
height = screen.get_height()

clock = pygame.time.Clock()
theScreen = Screen(screen.get_width(), screen.get_height())
sounds = music()

userHealth = healthBar(20, 20, 200, 30, 100, "left")
enemyHealth = healthBar(width - 20 - 200, 20, 200, 30, 100, "right")
damageStats = damage(50, 10)

cat1 = button(fuschia, 400, 200, 30, 30, " " + ws['C2'].value)
cat2 = button(fuschia, 650, 200, 30, 30, " " + ws['C7'].value)
cat3 = button(fuschia, 900, 200, 30, 30, " " + ws['C12'].value)
cat4 = button(fuschia, 450, 350, 30, 30, " " + ws['C17'].value)
cat5 = button(fuschia, 800, 350, 30, 30, " General")

a = button(color_light, 200, 175, 25, 25)
b = button(color_light, 200, 225, 25, 25)
c = button(color_light, 200, 275, 25, 25)
d = button(color_light, 200, 325, 25, 25)

chooseCat = text(fuschia, (theScreen.width * 0.25), (theScreen.height * 0.5 / 12), (theScreen.width * 0.5),
                            (theScreen.height * 1 / 10),
                            100, ["Choose a category:"])

#start menu buttons
title = text(black, (width * 0.25), (height * 0 / 12), (width * 0.5), (height * 2 / 10), 150,
                    ["PYTHON PALS"])
startButton = button(color_light, (width * 0.25), (height * 3 / 12), (width * 0.5), (height * 2 / 10),
                                       "Start")
optionsButton = button(color_light, (width * 0.25), (height * 6 / 12), (width * 0.5), (height * 2 / 10),
                              "Options")
quitButton = button(color_light, (width * 0.25), (height * 9 / 12), (width * 0.5), (height * 2 / 10),
                                      "Quit")
#options menu buttons
optionsHeader = text(black, (width * 0.25), (height * 0 / 12), (width * 0.5), (height * 2 / 10), 150,
                               ["OPTIONS"])
choose_q_deck = button(color_light, (width * 0.25), (height * 3 / 12), (width * 0.5), (height * 2 / 10),
                                      "Choose Question Deck")
change_sound_settings = button(color_light, (width * 0.25), (height * 6 / 12), (width * 0.5), (height * 2 / 10),
                                      "Change Sound Settings")
back_to_main = button(color_light, (width * 0.25), (height * 9 / 12), (width * 0.5), (height * 2 / 10),
                                      "Back to Main Menu")

#options -> return to options
back_to_options = button(color_light, (width * 0.25), (height * 9 / 12), (width * 0.5), (height * 2 / 10),
                                      "Back to Options Menu")

#options -> choose question deck
q_deck_header = text(black, (width * 0.25), (height * 0 / 12), (width * 0.5), (height * 2 / 10), 150,
                               ["QUESTION DECKS"])
changeCapitals = button(color_light, (width * 0.25), (height * 3 / 14), (width * 0.5), (height * 2 / 14),
                                      "Use Capitals Deck")
changePython = button(color_light, (width * 0.25), (height * 5.5 / 14), (width * 0.5), (height * 2 / 14),
                                      "Use Python Deck")
changeHistory = button(color_light, (width * 0.25), (height * 8 / 14), (width * 0.5), (height * 2 / 14),
                                      "Use History Deck")

#options -> change sound settings
sound_header = text(black, (width * 0.25), (height * 0 / 12), (width * 0.5), (height * 2 / 10), 150,
                               ["SOUND SETTINGS"])
mute_button = button(color_light, (width * 0.24), (height * 2.5 / 12), (width * 0.25), (height * 1 / 14),
                                      "Mute")
unmute_button = button(color_light, (width * 0.51), (height * 2.5 / 12), (width * 0.25), (height * 1 / 14),
                                      "Unmute")
music_label = text(black,(width * 0.25), (height * 3.25 / 12), (width * 0.5), (height * 2 / 14), 75,
                               ["Choose lobby music:"])
music_choice_1 = button(color_light, (width * 0.235), (height * 4.8 / 12), (width * 0.167), (height * 2 / 14),
                                      "8-bit")
music_choice_2 = button(color_light, (width * 0.4175), (height * 4.8 / 12), (width * 0.167), (height * 2 / 14),
                                      "Elf")
music_choice_3 = button(color_light, (width * 0.6), (height * 4.8 / 12), (width * 0.167), (height * 2 / 14),
                                      "Latin")
music_choice_4 = button(color_light, (width * 0.24), (height * 6.8 / 12), (width * 0.25), (height * 2 / 14),
                                      "Sinatra")
music_choice_5 = button(color_light, (width * 0.51), (height * 6.8 / 12), (width * 0.25), (height * 2 / 14),
                                      "Rick n' Roll")


def resize(startWidth, startHeight):
    #screen
    theScreen.width = startWidth
    theScreen.height = startHeight
    chooseCat.modify((theScreen.width * 0.25), (theScreen.height * 0 / 12), (theScreen.width * 0.5),
                     (theScreen.height * 1 / 10))

    #battle category/answer buttons
    cat1.modify(0.25 * startWidth, 0.22 * startHeight, startWidth * 0.025, startHeight * 0.025)
    cat2.modify(0.41 * startWidth, 0.22 * startHeight, startWidth * 0.025, startHeight * 0.025)
    cat3.modify(0.56 * startWidth, 0.22 * startHeight, startWidth * 0.025, startHeight * 0.025)
    cat4.modify(0.25 * startWidth, 0.4 * startHeight, startWidth * 0.025, startHeight * 0.025)
    cat5.modify(0.5 * startWidth, 0.4 * startHeight, startWidth * 0.025, startHeight * 0.025)

    #start menu
    title.modify((startWidth * 0.25), (startHeight * 0 / 12), (startWidth * 0.5), (startHeight * 2 / 10))
    startButton.modify((startWidth * 0.25), (startHeight * 3 / 12), (startWidth * 0.5), (startHeight * 2 / 10))
    optionsButton.modify((startWidth * 0.25), (startHeight * 6 / 12), (startWidth * 0.5), (startHeight * 2 / 10))
    quitButton.modify((startWidth * 0.25), (startHeight * 9 / 12), (startWidth * 0.5), (startHeight * 2 / 10))

    #options menu
    optionsHeader.modify((startWidth * 0.25), (startHeight * 0 / 12), (startWidth * 0.5), (startHeight * 2 / 10))
    choose_q_deck.modify((startWidth * 0.25), (startHeight * 3 / 12), (startWidth * 0.5), (startHeight * 2 / 10))
    change_sound_settings.modify((startWidth * 0.25), (startHeight * 6 / 12), (startWidth * 0.5), (startHeight * 2 / 10))
    back_to_main.modify((startWidth * 0.25), (startHeight * 9 / 12), (startWidth * 0.5), (startHeight * 2 / 10))

    #options -> submenus
    back_to_options.modify((startWidth * 0.25), (startHeight * 9 / 12), (startWidth * 0.5), (startHeight * 2 / 10))

    #options -> choose question deck menu
    q_deck_header.modify((startWidth * 0.25), (startHeight * 0 / 12), (startWidth * 0.5), (startHeight * 2 / 10))
    changeCapitals.modify((startWidth * 0.25), (startHeight * 3 / 14), (startWidth * 0.5), (startHeight * 2 / 14))
    changePython.modify((startWidth * 0.25), (startHeight * 5.5 / 14), (startWidth * 0.5), (startHeight * 2 / 14))
    changeHistory.modify((startWidth * 0.25), (startHeight * 8 / 14), (startWidth * 0.5), (startHeight * 2 / 14))

    #options -> change sound settings menu
    sound_header.modify((startWidth * 0.25), (startHeight * 0 / 12), (startWidth * 0.5), (startHeight * 2 / 10))
    mute_button.modify((startWidth * 0.24), (startHeight * 2.5 / 12), (startWidth * 0.25), (startHeight * 1 / 14))
    unmute_button.modify((startWidth * 0.51), (startHeight * 2.5 / 12), (startWidth * 0.25), (startHeight * 1 / 14))
    music_label.modify((startWidth * 0.25), (startHeight * 3.25 / 12), (startWidth * 0.5), (startHeight * 2 / 14))
    music_choice_1.modify((startWidth * 0.235), (startHeight * 4.8 / 12), (startWidth * 0.167), (startHeight * 2 / 14))
    music_choice_2.modify((startWidth * 0.4175), (startHeight * 4.8 / 12), (startWidth * 0.167), (startHeight * 2 / 14))
    music_choice_3.modify((startWidth * 0.6), (startHeight * 4.8 / 12), (startWidth * 0.167), (startHeight * 2 / 14))
    music_choice_4.modify((startWidth * 0.24), (startHeight * 6.8 / 12), (startWidth * 0.25), (startHeight * 2 / 14))
    music_choice_5.modify((startWidth * 0.51), (startHeight * 6.8 / 12), (startWidth * 0.25), (startHeight * 2 / 14))

def reDrawStartWindow():
    screen.fill(gold)
    title.draw(screen, int(theScreen.width*0.1), True)
    startButton.draw(screen, int(theScreen.width*0.05), black, True)
    optionsButton.draw(screen, int(theScreen.width*0.05), black, True)
    quitButton.draw(screen, int(theScreen.width*0.05), black, True)

def startscreen():
    resize(theScreen.width, theScreen.height)
    reDrawStartWindow()
    while True:
        # reDrawStartWindow()
        pygame.display.update()
        for ev in pygame.event.get():
            pos = pygame.mouse.get_pos()
            if ev.type == pygame.QUIT:
                pygame.quit()
                quit()

            if ev.type == pygame.MOUSEBUTTONDOWN:
                if startButton.isOver(pos):
                    # print("clicked start button")
                    return "start"
                elif optionsButton.isOver(pos):
                    # print("clicked options button")
                    return "options"
                elif quitButton.isOver(pos):
                    # print("clicked quit button")
                    return "quit"

            if ev.type == pygame.VIDEORESIZE:
                if (ev.w != theScreen.width):
                    theScreen.width = ev.w
                    theScreen.height = int(theScreen.width * 900 / 1600)
                else:
                    theScreen.height = ev.h
                    theScreen.width = int(theScreen.height * 1600 / 900)

                pygame.display.set_mode((theScreen.width, theScreen.height), pygame.RESIZABLE)
                resize(theScreen.width, theScreen.height)
                reDrawStartWindow()

        pygame.display.update()

def drawCategories(color, text_color):
    chooseCat = text(color, (theScreen.width * 0.25), (theScreen.height * 0.5 / 12), (theScreen.width * 0.5), (theScreen.height * 1 / 10), 100,
                     ["Choose a category:"])
    cat1 = button(color, theScreen.width * 0.25, theScreen.height * 0.22, theScreen.width * 0.025, theScreen.height * 0.025, " " + ws['C2'].value)
    cat2 = button(color, theScreen.width * 0.41, theScreen.height * 0.22, theScreen.width * 0.025, theScreen.height * 0.025, " " + ws['C7'].value)
    cat3 = button(color, theScreen.width * 0.56, theScreen.height * 0.22, theScreen.width * 0.025, theScreen.height * 0.025, " " + ws['C12'].value)
    cat4 = button(color, theScreen.width * 0.25, theScreen.height * 0.4, theScreen.width * 0.025, theScreen.height * 0.025, " " + ws['C17'].value)
    cat5 = button(color, theScreen.width * 0.5, theScreen.height * 0.4, theScreen.width * 0.025, theScreen.height * 0.025, " General")

    chooseCat.draw(screen, int(theScreen.width * 0.07), True)
    cat1.draw(screen, int(theScreen.width * 0.02), text_color, False)
    cat2.draw(screen, int(theScreen.width * 0.02), text_color, False)
    cat3.draw(screen, int(theScreen.width * 0.02), text_color, False)
    cat4.draw(screen, int(theScreen.width * 0.02), text_color, False)
    cat5.draw(screen, int(theScreen.width * 0.02), text_color, False)

def chooseCategory(playerGroup, enemyGroup,level,color,text_color,myPlayer,myEnemy):
    drawCategories(color, text_color)
    while True:
        for ev in pygame.event.get():
            pos = pygame.mouse.get_pos()
            if ev.type == pygame.QUIT:
                pygame.quit()
                quit()

            if ev.type == pygame.MOUSEBUTTONDOWN:
                if cat1.isOver(pos):
                    print(1)
                    return 1
                elif cat2.isOver(pos):
                    print(2)
                    return 2
                elif cat3.isOver(pos):
                    print(3)
                    return 3
                elif cat4.isOver(pos):
                    print(4)
                    return 4
                elif cat5.isOver(pos):
                    print(5)
                    return 5

            if ev.type == pygame.VIDEORESIZE:
                if (ev.w != theScreen.width):
                    theScreen.width = ev.w
                    theScreen.height = int(theScreen.width * 900 / 1600)
                else:
                    theScreen.height = ev.h
                    theScreen.width = int(theScreen.height * 1600 / 900)

                pygame.display.set_mode((theScreen.width, theScreen.height), pygame.RESIZABLE)
                resizeBattle(theScreen.width, theScreen.height)
                myPlayer.resize()
                playerGroup = pygame.sprite.Group(myPlayer)
                myEnemy.resize()
                enemyGroup = pygame.sprite.Group(myEnemy)
                drawBattle(playerGroup, enemyGroup, theScreen.width, theScreen.height,level)
                drawCategories(color, text_color)

def resizeBattle(startWidth, startHeight):
    userHealth.modify(0.02 * startWidth, 0.02 * startHeight, startWidth * 0.1, startHeight * 0.03)
    enemyHealth.modify(0.88 * startWidth, 0.02 * startHeight, startWidth * 0.1, startHeight * 0.03)

    a.modify(0.125 * startWidth, 0.19 * startHeight, startWidth * 0.021, startHeight * 0.021)
    b.modify(0.125 * startWidth, 0.25 * startHeight, startWidth * 0.021, startHeight * 0.021)
    c.modify(0.125 * startWidth, 0.305 * startHeight, startWidth * 0.021, startHeight * 0.021)
    d.modify(0.125 * startWidth, 0.36 * startHeight, startWidth * 0.021, startHeight * 0.021)

    chooseCat.modify(0.25 * startWidth, 1/24 * startHeight, startWidth * 0.5, startHeight * 0.1)

    cat1.modify(0.25 * startWidth, 0.22 * startHeight, startWidth * 0.025, startHeight * 0.025)
    cat2.modify(0.41 * startWidth, 0.22 * startHeight, startWidth * 0.025, startHeight * 0.025)
    cat3.modify(0.56 * startWidth, 0.22 * startHeight, startWidth * 0.025, startHeight * 0.025)
    cat4.modify(0.25 * startWidth, 0.4 * startHeight, startWidth * 0.025, startHeight * 0.025)
    cat5.modify(0.5 * startWidth, 0.4 * startHeight, startWidth * 0.025, startHeight * 0.025)

    pygame.display.update()

def drawAnswers(text_color):
    a.draw(screen, int(theScreen.width * 0.02), text_color, False)
    b.draw(screen, int(theScreen.width * 0.02), text_color, False)
    c.draw(screen, int(theScreen.width * 0.02), text_color, False)
    d.draw(screen, int(theScreen.width * 0.02), text_color, False)

def chooseAnswer(playerGroup, enemyGroup, question ,level,q,A,B,C,D, text_color,myPlayer,myEnemy):
    drawAnswers(text_color)
    while True:
        for ev in pygame.event.get():
            pos = pygame.mouse.get_pos()
            if ev.type == pygame.QUIT:
                pygame.quit()
                quit()

            if ev.type == pygame.MOUSEBUTTONDOWN:
                if a.isOver(pos):
                    return 'a'
                elif b.isOver(pos):
                    return 'b'
                elif c.isOver(pos):
                    return 'c'
                elif d.isOver(pos):
                    return 'd'

            if ev.type == pygame.VIDEORESIZE:
                if (ev.w != theScreen.width):
                    theScreen.width = ev.w
                    theScreen.height = int(theScreen.width * 900 / 1600)
                else:
                    theScreen.height = ev.h
                    theScreen.width = int(theScreen.height * 1600 / 900)

                pygame.display.set_mode((theScreen.width, theScreen.height), pygame.RESIZABLE)
                resizeBattle(theScreen.width, theScreen.height)
                myPlayer.resize()
                playerGroup = pygame.sprite.Group(myPlayer)
                myEnemy.resize()
                enemyGroup = pygame.sprite.Group(myEnemy)
                drawBattle(playerGroup, enemyGroup, theScreen.width, theScreen.height,level)

                q.modify(theScreen.width * 0.25, theScreen.height * 1/9, 50, 50)
                A.modify(a.x, a.y, 50, 50)
                B.modify(b.x, b.y, 50, 50)
                C.modify(c.x, c.y, 50, 50)
                D.modify(d.x, d.y, 50, 50)

                q.draw(screen, int(theScreen.width * 0.02), False)
                A.draw(screen, int(theScreen.width * 0.015), False)
                B.draw(screen, int(theScreen.width * 0.015), False)
                C.draw(screen, int(theScreen.width * 0.015), False)
                D.draw(screen, int(theScreen.width * 0.015), False)
                drawAnswers(text_color)

def drawBattle(playerGroup, enemyGroup, w, h,level):
    screen.fill((0,0,0))
    #pygame.draw.rect(screen, (0, 0, 0), (0, 0, theScreen.width, theScreen.height * 0.1))
    # bg = pygame.image.load("jungleBackground.jpg")
    drawBackground(playerGroup,enemyGroup,w,h,level)
    # bg = pygame.transform.rotozoom(bg, 0, theScreen.width/ 1380)
    # screen.blit(bg,(0,0))
    #pygame.draw.rect(screen, (0, 0, 0), (0, 0, theScreen.width, theScreen.height * 0.07))
    playerGroup.draw(screen)
    enemyGroup.draw(screen)
    enemyHealth.draw(screen)
    userHealth.draw(screen)
    theScreen.width = w
    theScreen.height = h
    pygame.display.update()

#correctOrWrong is optional parameter, only use if if an animation that has has answer stuff
def drawBackground(playerGroup, enemyGroup, width, height, level, correctOrWrong=""):
    # pygame.display.update()
    # screen.fill(white)
    #pygame.draw.rect(screen, (0, 0, 0), (0, 0, theScreen.width, theScreen.height * 0.07))
    if level == 1:
        bg = pygame.image.load("jungleBackground.jpg")
        bg = pygame.transform.rotozoom(bg, 0, theScreen.width / 1380)
    elif level == 2:
        bg = pygame.image.load("goldmine.jpg")
        bg = pygame.transform.rotozoom(bg, 0, theScreen.width / 1400)
    else:
        bg = pygame.image.load("ocean.jpg")
        bg = pygame.transform.rotozoom(bg, 0, theScreen.width / 1850)
    screen.blit(bg,(0,0))
    if correctOrWrong != "":
        if correctOrWrong == "Correct":
            #pygame.mixer.Sound("correct.wav").play(0)
            display = text(black, theScreen.width * 0.25, theScreen.height * 0.2, theScreen.width * 0.5,
                                  theScreen.height * 0.4, int(theScreen.width * 0.02), ["Correct"])
            pygame.draw.ellipse(screen, (0, 200, 0), (theScreen.width * 0.35, theScreen.height * 0.3, theScreen.width * 0.3, theScreen.height * 0.2))
        if correctOrWrong == "Wrong":
            #pygame.mixer.Sound("incorrect.wav").play(0)
            display = text(black, theScreen.width * 0.25, theScreen.height * 0.2, theScreen.width * 0.5,
                                  theScreen.height * 0.4, int(theScreen.width * 0.02), ["Incorrect"])
            pygame.draw.ellipse(screen, (255, 0, 0), (theScreen.width * 0.35, theScreen.height * 0.3, theScreen.width * 0.3, theScreen.height * 0.2))
        display.draw(screen, int(theScreen.width * 1/16), True)
    # playerGroup.draw(screen)
    # enemyGroup.draw(screen)
    enemyHealth.draw(screen)
    userHealth.draw(screen)
    # pygame.display.update()

def animationController(playerGroup, enemyGroup, width, height, clock, level, animation):
    if(animation == "snake attack"):
        for i in range(0,4):
            drawBackground(playerGroup, enemyGroup, width, height,level, "Correct")
            playerGroup.update()
            playerGroup.draw(screen)
            enemyGroup.draw(screen)
            pygame.display.update()
            clock.tick(5)
        drawBackground(playerGroup, enemyGroup, width, height,level, "Correct")
        playerGroup.update()

        playerGroup.update()
        playerGroup.update()
        playerGroup.draw(screen)
        enemyGroup.draw(screen)
        pygame.display.update()

    if(animation == "coffee hurt"):
        drawBackground(playerGroup, enemyGroup, width, height,level, "Correct")
        enemyGroup.update()
        enemyGroup.update()
        enemyGroup.update()
        enemyGroup.draw(screen)
        playerGroup.draw(screen)
        pygame.display.update()
        clock.tick(3)

        drawBackground(playerGroup, enemyGroup, width, height,level, "Correct")
        enemyGroup.update()
        enemyGroup.update()
        enemyGroup.update()
        enemyGroup.draw(screen)
        playerGroup.draw(screen)
        pygame.display.update()
        
    if(animation == "coffee attack"):
        drawBackground(playerGroup, enemyGroup, width, height,level, "Wrong")
        enemyGroup.update()
        enemyGroup.draw(screen)
        playerGroup.draw(screen)
        pygame.display.update()
        clock.tick(3)

        drawBackground(playerGroup, enemyGroup, width, height,level, "Wrong")
        enemyGroup.update()
        enemyGroup.draw(screen)
        playerGroup.draw(screen)
        pygame.display.update()
        clock.tick(3)

        drawBackground(playerGroup, enemyGroup, width, height,level, "Wrong")
        enemyGroup.update()
        enemyGroup.update()
        enemyGroup.update()
        enemyGroup.update()
        enemyGroup.draw(screen)
        playerGroup.draw(screen)
        pygame.display.update()

    if(animation == "snake hurt"):
        drawBackground(playerGroup, enemyGroup, width, height,level, "Wrong")
        for i in range(0, 4):
            playerGroup.update()
        playerGroup.update()
        playerGroup.draw(screen)
        enemyGroup.draw(screen)
        pygame.display.update()
        clock.tick(3)
        drawBackground(playerGroup, enemyGroup, width, height,level, "Wrong")
        playerGroup.update()
        playerGroup.update()
        playerGroup.draw(screen)
        enemyGroup.draw(screen)
        pygame.display.update()

    if (animation == "coffee break"):
        clock.tick(2)
        for i in range (0,6):
            playerGroup.update()
        for i in range (0,3):
            enemyGroup.update()
        drawBackground(playerGroup, enemyGroup, width, height,level)
        enemyGroup.update()
        enemyGroup.draw(screen)
        playerGroup.draw(screen)
        pygame.display.update()
        clock.tick(2)
        drawBackground(playerGroup, enemyGroup, width, height,level)
        enemyGroup.update()
        enemyGroup.draw(screen)
        playerGroup.draw(screen)
        pygame.display.update()
        clock.tick(2)
    if(animation == "ruby attack"):
        for i in range(0, 7):
            drawBackground(playerGroup, enemyGroup, width, height,level, "Wrong")
            enemyGroup.update()
            playerGroup.draw(screen)
            enemyGroup.draw(screen)
            pygame.display.update()
            clock.tick(5)

        drawBackground(playerGroup, enemyGroup, width, height,level, "Wrong")
        enemyGroup.update()
        enemyGroup.update()
        playerGroup.draw(screen)
        enemyGroup.draw(screen)
        pygame.display.update()

    if (animation == "ruby hurt"):
        for i in range(0, 7):
            enemyGroup.update()

        drawBackground(playerGroup, enemyGroup, width, height,level, "Correct")
        enemyGroup.update()
        playerGroup.draw(screen)
        enemyGroup.draw(screen)
        pygame.display.update()
        clock.tick(3)

        drawBackground(playerGroup, enemyGroup, width, height,level, "Correct")
        enemyGroup.update()
        playerGroup.draw(screen)
        enemyGroup.draw(screen)
        pygame.display.update()

    if (animation == "eye attack"):
        for i in range(0, 2):
            drawBackground(playerGroup, enemyGroup, width, height, level, "Wrong")
            enemyGroup.update()
            playerGroup.draw(screen)
            enemyGroup.draw(screen)
            pygame.display.update()
            clock.tick(2)

        drawBackground(playerGroup, enemyGroup, width, height, level, "Wrong")
        enemyGroup.update()
        enemyGroup.update()
        playerGroup.draw(screen)
        enemyGroup.draw(screen)
        pygame.display.update()

    if (animation == "eye hurt"):
        for i in range(0, 2):
            enemyGroup.update()

        drawBackground(playerGroup, enemyGroup, width, height, level, "Correct")
        enemyGroup.update()
        playerGroup.draw(screen)
        enemyGroup.draw(screen)
        pygame.display.update()
        clock.tick(3)

        drawBackground(playerGroup, enemyGroup, width, height, level, "Correct")
        enemyGroup.update()
        playerGroup.draw(screen)
        enemyGroup.draw(screen)
        pygame.display.update()

def theBattle(level, play):
    health = 100
    enemy_health = 100
    userHealth.set_health(health)
    enemyHealth.set_health(enemy_health)

    resizeBattle(theScreen.width, theScreen.height)
    clock = pygame.time.Clock()

    myPlayer = Player()
    playerGroup = pygame.sprite.Group(myPlayer)
    # playerGroup.draw(screen)

    cat_button_color = white
    ans_button_color = white
    text_color = white
    # level = 3
    if level == 1:
        myEnemy = Coffee()
        if play:
            pygame.mixer.music.load('jazz.mp3')
            pygame.mixer.music.play(-1)
        text_color = black
        cat_button_color = (215,132,37)#(122,135,117)
        ans_button_color = (217,125,19)
        box_color = (77, 160, 83, 130)
    elif level == 2:
        myEnemy = Ruby()
        if play:
            pygame.mixer.music.load('funke.mp3')
            pygame.mixer.music.play(-1)
        cat_button_color = (240,208,79)
        text_color = (204,203,194)
        ans_button_color = color_dark
        box_color = (0,0,0,0)
    else:
        myEnemy = Eye()
        if play:
            pygame.mixer.music.load('bluth.wav')
            pygame.mixer.music.play(-1)
        cat_button_color = (12,37,99)
        text_color = (170,171,173)
        ans_button_color = (1,26,41)
        box_color = (0,0,0,0)

    enemyGroup = pygame.sprite.Group(myEnemy)
    # enemyGroup.draw(screen)
    pygame.display.update()

    battle = True
    while battle:
        # myPlayer.resize()
        # playerGroup = pygame.sprite.Group(myPlayer)
        clock.tick(27)
        userHealth.modify(0.02 * theScreen.width, 0.02 * theScreen.height, theScreen.width * 0.1,
                          theScreen.height * 0.03)
        enemyHealth.modify(0.88 * theScreen.width, 0.02 * theScreen.height, theScreen.width * 0.1,
                           theScreen.height * 0.03)
        # enemyGroup.draw(screen)
        # playerGroup.draw(screen)
        drawBattle(playerGroup, enemyGroup, theScreen.width, theScreen.height,level)
        choose = chooseCategory(playerGroup, enemyGroup, level, cat_button_color, text_color,myPlayer,myEnemy)
        print(choose)
        questionNumber = load_question(choose)
        question = get_question(questionNumber, ws)

        drawBattle(playerGroup, enemyGroup, theScreen.width, theScreen.height,level)

        #ellipse
        pygame.draw.ellipse(screen, box_color, (0, int(theScreen.height * 0.1), int(theScreen.width), int(theScreen.height * 0.1)))

        q = text(text_color, theScreen.width * 0.25, theScreen.height * 1/9, 50, 50, 40, question[0].splitlines())
        A = text(text_color, a.x, a.y, 50, 50, 40, question[1].splitlines())
        B = text(text_color, b.x, b.y, 50, 50, 40, question[2].splitlines())
        C = text(text_color, c.x, c.y, 50, 50, 40, question[3].splitlines())
        D = text(text_color, d.x, d.y, 50, 50, 40, question[4].splitlines())

        #question box
        #pygame.gfxdraw.box(screen, pygame.Rect(theScreen.width * 0.1, theScreen.height * 0.15,
                                               #600, 300), box_color)

        q.draw(screen, int(theScreen.width*0.02), False)
        A.draw(screen, int(theScreen.width*0.015), False)
        B.draw(screen, int(theScreen.width*0.015), False)
        C.draw(screen, int(theScreen.width*0.015), False)
        D.draw(screen, int(theScreen.width*0.015), False)

        guess = chooseAnswer(playerGroup, enemyGroup, question,level,q,A,B,C,D, ans_button_color,myPlayer,myEnemy)
        is_correct = get_result(guess, question[5])

        if is_correct:
            enemy_health = enemy_health - damageStats.userDamage
            enemyHealth.set_health(enemy_health)
            enemyHealth.draw(screen)

            clock.tick(5)
            animationController(playerGroup,enemyGroup,theScreen.width,theScreen.height, clock,level, "snake attack")
            if level == 1:
                animationController(playerGroup, enemyGroup, theScreen.width, theScreen.height, clock,level, "coffee hurt")
            elif level == 2:
                animationController(playerGroup, enemyGroup, theScreen.width, theScreen.height, clock,level, "ruby hurt")
            else:
                animationController(playerGroup, enemyGroup, theScreen.width, theScreen.height, clock, level, "eye hurt")
        else:
            health = health - damageStats.enemyDamage
            userHealth.set_health(health)
            userHealth.draw(screen)

            clock.tick(3)
            if level == 1:
                animationController(playerGroup, enemyGroup, theScreen.width, theScreen.height, clock,level, "coffee attack")
            elif level == 2:
                animationController(playerGroup, enemyGroup, theScreen.width, theScreen.height, clock,level, "ruby attack")
            else:
                animationController(playerGroup, enemyGroup, theScreen.width, theScreen.height, clock, level, "eye attack")
            animationController(playerGroup, enemyGroup, theScreen.width, theScreen.height, clock,level, "snake hurt")

        if health <= 0:
            battle = False
            return "lose"

        if enemy_health <= 0:
            battle = False
            if level == 1:
                animationController(playerGroup, enemyGroup, theScreen.width, theScreen.height, clock,level, "coffee break")
            return "win"

def win(play):
    if play:
        pygame.mixer.music.load('victoire.mp3')
        pygame.mixer.music.play(-1)

    print("Victory!")

    screen.fill(purple) #gold
    fw = pygame.image.load("fireworks.png")
    fw = pygame.transform.rotozoom(fw, 0, theScreen.width / 820)
    screen.blit(fw, (theScreen.width * 0.2, theScreen.height * 0.25))
    snake = pygame.image.load("snake6.png")
    snake.convert()
    snake = pygame.transform.rotozoom(snake, 0, theScreen.height / 600)
    screen.blit(snake, (theScreen.width * 0.4, theScreen.height * 0.6))
    winScreen = text(black, (theScreen.width * 0.25), (theScreen.height * 0 / 12), (theScreen.width * 0.5),
                            (theScreen.height * 2 / 10), 150,
                            ["You win!"])
    winScreen.draw(screen, int(theScreen.width*0.1), True)
    clock.tick(2)
    pygame.display.update()
    pygame.time.delay(4000)

def lose(play):
    if play:
        pygame.mixer.music.load('defaite.mp3')
        pygame.mixer.music.play(-1)

    print("Defeat")
    screen.fill(blue) #color_light
    rain = pygame.image.load("rain.png")
    rain = pygame.transform.rotozoom(rain, 0, theScreen.width / 2000)
    screen.blit(rain, (theScreen.width * 0.4, theScreen.height * 0.25))

    snake = pygame.image.load("snake5.png")
    snake.convert()
    snake = pygame.transform.rotozoom(snake, 0, theScreen.height / 600)
    screen.blit(snake, (theScreen.width * 0.34, theScreen.height * 0.6))
    gameOver = text(indigo, (theScreen.width * 0.25), (theScreen.height * 0 / 12), (theScreen.width * 0.5), (theScreen.height * 2 / 10),
                           150,
                           ["You lose!"])
    gameOver.draw(screen, int(theScreen.width*0.1), True)
    clock.tick(2)
    pygame.display.update()
    pygame.time.delay(4000)

def reDrawOptionsWindow():
    screen.fill(purple)
    optionsHeader.draw(screen, int(theScreen.width*0.1), True)
    choose_q_deck.draw(screen, int(theScreen.width*0.04), black, True)
    change_sound_settings.draw(screen, int(theScreen.width*0.04), black, True)
    back_to_main.draw(screen, int(theScreen.width*0.04), black, True)
    pygame.display.update()

def options():
    reDrawOptionsWindow()
    opt = True
    while opt:
        for ev in pygame.event.get():
            pos = pygame.mouse.get_pos()
            if ev.type == pygame.QUIT:
                pygame.quit()
                quit()

            if ev.type == pygame.MOUSEBUTTONDOWN:
                if back_to_main.isOver(pos):
                    opt = False
                if choose_q_deck.isOver(pos):
                    q_deck_menu()
                    opt = False
                if change_sound_settings.isOver(pos):
                    sound_settings_menu()
                    #return sound_settings_menu()
                    opt = False

            if ev.type == pygame.VIDEORESIZE:
                if (ev.w != theScreen.width):
                    theScreen.width = ev.w
                    theScreen.height = int(theScreen.width * 900 / 1600)
                else:
                    theScreen.height = ev.h
                    theScreen.width = int(theScreen.height * 1600 / 900)
                pygame.display.set_mode((theScreen.width, theScreen.height), pygame.RESIZABLE)
                resize(theScreen.width, theScreen.height)
                reDrawOptionsWindow()

#options -> choose question deck
def redraw_q_deck_window():
    screen.fill(blue)
    q_deck_header.draw(screen, int(theScreen.width*0.1), True)
    changeCapitals.draw(screen, int(theScreen.width*0.04), black, True)
    changePython.draw(screen, int(theScreen.width*0.04), black, True)
    changeHistory.draw(screen, int(theScreen.width*0.04), black, True)
    back_to_options.draw(screen, int(theScreen.width*0.04), black, True)
    pygame.display.update()

def q_deck_menu():
    redraw_q_deck_window()
    opt = True
    while opt:
        for ev in pygame.event.get():
            pos = pygame.mouse.get_pos()
            if ev.type == pygame.QUIT:
                pygame.quit()
                quit()

            if ev.type == pygame.MOUSEBUTTONDOWN:
                if back_to_options.isOver(pos):
                    options()
                    opt = False
                if changeCapitals.isOver(pos):
                    globals()['ws'] = changeQuestionDeck("python_questions_capitals.xlsx", ws)
                    options()
                    opt = False
                if changePython.isOver(pos):
                    globals()['ws'] = changeQuestionDeck("python_questions.xlsx", ws)
                    options()
                    opt = False
                if changeHistory.isOver(pos):
                    globals()['ws'] = changeQuestionDeck("python_questions_timeline.xlsx", ws)
                    options()
                    opt = False

            if ev.type == pygame.VIDEORESIZE:
                if (ev.w != theScreen.width):
                    theScreen.width = ev.w
                    theScreen.height = int(theScreen.width * 900 / 1600)
                else:
                    theScreen.height = ev.h
                    theScreen.width = int(theScreen.height * 1600 / 900)

                pygame.display.set_mode((theScreen.width, theScreen.height), pygame.RESIZABLE)
                resize(theScreen.width, theScreen.height)
                redraw_q_deck_window()

#options -> change sound settings
def redraw_sound_window():
    screen.fill(blue)
    sound_header.draw(screen, int(theScreen.width*0.1), True)
    unmute_button.draw(screen, int(theScreen.width*0.032), black, True)
    mute_button.draw(screen, int(theScreen.width*0.032), black, True)
    music_label.draw(screen, int(theScreen.width*0.05), True)
    music_choice_1.draw(screen, int(theScreen.width*0.04), black, True)
    music_choice_2.draw(screen, int(theScreen.width * 0.04), black, True)
    music_choice_3.draw(screen, int(theScreen.width * 0.04), black, True)
    music_choice_4.draw(screen, int(theScreen.width * 0.04), black, True)
    music_choice_5.draw(screen, int(theScreen.width * 0.04), black, True)
    back_to_options.draw(screen, int(theScreen.width*0.04), black, True)
    pygame.display.update()

def sound_settings_menu():
    redraw_sound_window()
    opt = True
    while opt:
        for ev in pygame.event.get():
            pos = pygame.mouse.get_pos()
            if ev.type == pygame.QUIT:
                pygame.quit()
                quit()

            if ev.type == pygame.MOUSEBUTTONDOWN:
                if back_to_options.isOver(pos):
                    options()
                    opt = False
                if music_choice_1.isOver(pos):
                    pygame.mixer.music.load('8bit_game.mp3')
                    pygame.mixer.music.play(-1)
                    #sound_settings_menu()
                    #return "8bit"
                    sounds.change("8bit")
                    opt = True
                if music_choice_2.isOver(pos):
                    pygame.mixer.music.load('idle.wav')
                    pygame.mixer.music.play(-1)
                    #sound_settings_menu()
                    #return "elf"
                    sounds.change("elf")
                    opt = True
                if music_choice_3.isOver(pos):
                    pygame.mixer.music.load('latin.wav')
                    pygame.mixer.music.play(-1)
                    #sound_settings_menu()
                    #return "latin"
                    sounds.change("latin")
                    opt = True
                if music_choice_4.isOver(pos):
                    pygame.mixer.music.load('elevator_music_16bit.wav')
                    pygame.mixer.music.play(-1)
                    #sound_settings_menu()
                    #return "sinatra"
                    sounds.change("sinatra")
                    opt = True
                if music_choice_5.isOver(pos):
                    pygame.mixer.music.load('lobby_rick_roll.wav')
                    pygame.mixer.music.play(-1)
                    #sound_settings_menu()
                    #return "rick_roll"
                    sounds.change("rick_roll")
                    opt = True
                if unmute_button.isOver(pos):
                    pygame.mixer.music.unpause()
                    #sound_settings_menu()
                    sounds.unmute()
                    opt = True
                    #return False #"unmuted"
                if mute_button.isOver(pos):
                    pygame.mixer.music.pause()
                    #sound_settings_menu()
                    sounds.mute()
                    opt = True
                    #return True #"muted"

            if ev.type == pygame.VIDEORESIZE:
                if (ev.w != theScreen.width):
                    theScreen.width = ev.w
                    theScreen.height = int(theScreen.width * 900 / 1600)
                else:
                    theScreen.height = ev.h
                    theScreen.width = int(theScreen.height * 1600 / 900)

                pygame.display.set_mode((theScreen.width, theScreen.height), pygame.RESIZABLE)
                resize(theScreen.width, theScreen.height)
                redraw_sound_window()

def levelChange(screen, level, Enemy, enemyImg, sound):
    lev = text(black, theScreen.width * 0.25, theScreen.height * 0.1, theScreen.width * 0.5,
                        theScreen.height * 0.2, 200, ["LEVEL " + str(level)])
    if level == 1:
        enemy = text(black, theScreen.width * 0.25, theScreen.height * 0.35, theScreen.width * 0.5,
                     theScreen.height * 0.2, theScreen.width * 1/16, ["Warmup: Beat Java"])
    elif level == 2:
        enemy = text(black, theScreen.width * 0.25, theScreen.height * 0.35, theScreen.width * 0.5,
                     theScreen.height * 0.2, theScreen.width * 1 / 16, ["Your enemy is Ruby"])
    else:
        enemy = text(black, theScreen.width * 0.25, theScreen.height * 0.35, theScreen.width * 0.5,
                        theScreen.height * 0.2, theScreen.width * 0.01, ["Let's C if you can beat Eye"])
    screen.fill(gold)

    if sound:
        pygame.mixer.music.load("elevator_music_16bit.wav")
        pygame.mixer.music.play(0)

    enemyimg = pygame.image.load(enemyImg)
    enemyimg = pygame.transform.rotozoom(enemyimg, 0, theScreen.width / 2000)

    lev.draw(screen, int(theScreen.width*0.2), True)
    enemy.draw(screen, int(theScreen.width*0.07), True)
    if Enemy != "Eye":
        screen.blit(enemyimg, (theScreen.width * 0.4, theScreen.height * 0.55))
    else:
        screen.blit(enemyimg, (theScreen.width * 0.35, theScreen.height * 0.55))

    pygame.display.update()
    pygame.time.delay(2000)

def main():
    pygame.init()
    pygame.display.set_caption("Python Pals")

    enter_game = True
    restart_music = True

    while enter_game:
        if sounds.play and restart_music:
            if sounds.name == "8bit":
                pygame.mixer.music.load('8bit_game.mp3')
                pygame.mixer.music.play(-1)
            elif sounds.name == "latin":
                pygame.mixer.music.load('latin.wav')
                pygame.mixer.music.play(-1)
            elif sounds.name == "sinatra":
                pygame.mixer.music.load('elevator_music_16bit.wav')
                pygame.mixer.music.play(-1)
            elif sounds.name == "rick_roll":
                pygame.mixer.music.load('lobby_rick_roll.wav')
                pygame.mixer.music.play(-1)
            else: #sounds.name == "elf" #or ""
                pygame.mixer.music.load('idle.wav')
                pygame.mixer.music.play(-1)

        #start screen
        menuOption = startscreen()

        #go to start, options, or quit
        if menuOption == "start":
            damageStats.modify(50,10)
            levelChange(screen, 1, "Java", "coffee1.png", sounds.play)
            result = theBattle(1, sounds.play)
            restart_music = True
            if result == "win":
                levelChange(screen, 2, "Ruby", "Ruby_idle.png", sounds.play)
                damageStats.modify(25,25)
                result2 = theBattle(2, sounds.play)
                if result2 == "win":
                    levelChange(screen, 3, "Eye", "C1.png", sounds.play)
                    damageStats.modify(17,25)
                    result3 = theBattle(3, sounds.play)
                    if result3 == "win":
                        win(sounds.play)
                    else:
                        lose(sounds.play)
                else:
                    lose(sounds.play)
            else:
                lose(sounds.play)
        elif menuOption == "options":
            #sound = options()
            #sound_str = options()
            options()
            restart_music = False
            '''
            if sound_str == "unmuted":
                sound = True
            elif sound_str == "muted":
                sound = False
            #elif sound_str != "":
                #sound = True
            else: #sound_str == ""
                sound = True
                restart_music = False
            '''
        elif menuOption == "quit":
            enter_game = False
            #restart_music = False
    # pygame.mixer.music.stop()
    pygame.quit()

if __name__ == "__main__":
    main()
